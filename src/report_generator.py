import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime, timedelta
import os
from database_manager import db_manager
import calendar

class ReportGenerator:
    def __init__(self):
        # Carpeta fija de descargas solicitada por el usuario:
        # C:\Users\<usuario>\Documents\SISTEMAS\setups\nfc\sistema_asistencia\DESCARGAS
        # 1) Intentar leer una ruta persistida en la base local (configuraciones_local)
        base_download_dir = None
        try:
            c = db_manager.sqlite_connection.cursor()
            c.execute("SELECT valor FROM configuraciones_local WHERE clave = 'DESCARGAS_DIR'")
            row = c.fetchone()
            if row and row[0]:
                base_download_dir = os.path.expandvars(str(row[0]).strip())
        except Exception:
            base_download_dir = None

        # 2) Si no hay configuración persistida, usar variables de entorno o default y guardarla por primera vez
        if not base_download_dir:
            custom_download_dir = os.getenv('ASISTENCIA_DESCARGAS_DIR') or os.getenv('DESCARGAS_DIR')
            if custom_download_dir and isinstance(custom_download_dir, str) and custom_download_dir.strip():
                base_download_dir = os.path.expandvars(custom_download_dir.strip())
            else:
                user_docs = os.path.join(os.path.expanduser('~'), 'Documents')
                base_download_dir = os.path.join(
                    user_docs,
                    'SISTEMAS', 'setups', 'nfc', 'sistema_asistencia', 'DESCARGAS'
                )
            # Guardar persistido para futuros arranques en esta PC
            try:
                c = db_manager.sqlite_connection.cursor()
                c.execute("INSERT OR REPLACE INTO configuraciones_local (clave, valor) VALUES ('DESCARGAS_DIR', ?)", (base_download_dir,))
                db_manager.sqlite_connection.commit()
            except Exception:
                pass
        # Crear la carpeta si no existe
        os.makedirs(base_download_dir, exist_ok=True)

        # Unificar todas las exportaciones en la misma carpeta de DESCARGAS
        self.reports_dir = base_download_dir
        self.employee_reports_dir = base_download_dir
    
    def generate_daily_report(self, fecha=None, formato='both'):
        """Generar reporte diario en Excel y/o PDF"""
        if fecha is None:
            fecha = datetime.now().date()
        
        try:
            # Obtener datos del día
            data = self._get_daily_data(fecha)
            
            if not data:
                print(f"No hay registros para la fecha {fecha}")
                return None
            
            files_generated = []
            
            if formato in ['excel', 'both']:
                excel_file = self._generate_daily_excel(data, fecha)
                if excel_file:
                    files_generated.append(excel_file)
            
            if formato in ['pdf', 'both']:
                pdf_file = self._generate_daily_pdf(data, fecha)
                if pdf_file:
                    files_generated.append(pdf_file)
            
            return files_generated
            
        except Exception as e:
            print(f"Error generando reporte diario: {e}")
            return None
    
    def generate_monthly_report(self, year=None, month=None, formato='both'):
        """Generar reporte mensual en Excel y/o PDF"""
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month
        
        try:
            # Obtener datos del mes
            data = self._get_monthly_data(year, month)
            
            if not data:
                print(f"No hay registros para {month}/{year}")
                return None
            
            files_generated = []
            
            # Orden alfabético por apellidos (última palabra del nombre)
            # data viene como (empleado_id, nombre, fecha, ...)
            try:
                data = sorted(data, key=lambda r: (str(r[1]).split()[-1].upper(), str(r[1]).upper(), str(r[2])))
            except Exception:
                pass

            if formato in ['excel', 'both']:
                excel_file = self._generate_monthly_excel(data, year, month)
                if excel_file:
                    files_generated.append(excel_file)
            
            if formato in ['pdf', 'both']:
                pdf_file = self._generate_monthly_pdf(data, year, month)
                if pdf_file:
                    files_generated.append(pdf_file)
            
            return files_generated
            
        except Exception as e:
            print(f"Error generando reporte mensual: {e}")
            return None
    
    def generate_employee_report(self, empleado_id, year=None, month=None, formato='both'):
        """Generar reporte individual de empleado"""
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month
        
        try:
            # Obtener datos del empleado
            employee_data, attendance_data = self._get_employee_data(empleado_id, year, month)
            
            if not employee_data:
                print(f"Empleado con ID {empleado_id} no encontrado")
                return None
            
            files_generated = []
            
            if formato in ['excel', 'both']:
                excel_file = self._generate_employee_excel(employee_data, attendance_data, year, month)
                if excel_file:
                    files_generated.append(excel_file)
            
            if formato in ['pdf', 'both']:
                pdf_file = self._generate_employee_pdf(employee_data, attendance_data, year, month)
                if pdf_file:
                    files_generated.append(pdf_file)
            
            return files_generated
            
        except Exception as e:
            print(f"Error generando reporte de empleado: {e}")
            return None

    def generate_employee_daily_report(self, empleado_id, fecha=None, formato='both'):
        """Generar reporte diario de un empleado (PDF y/o Excel)."""
        try:
            if fecha is None:
                fecha = datetime.now().date()

            # Obtener datos del empleado y registros del día
            if db_manager.is_online():
                cursor = db_manager.pg_connection.cursor()
                cursor.execute("""
                    SELECT nombre_completo, cargo, rol, hora_entrada, hora_salida
                    FROM empleados WHERE id = %s AND activo = TRUE
                """, (empleado_id,))
                employee_data = cursor.fetchone()

                cursor.execute("""
                    SELECT fecha, hora_registro, tipo_movimiento, estado
                    FROM registros_asistencia 
                    WHERE empleado_id = %s AND fecha = %s
                    ORDER BY hora_registro
                """, (empleado_id, fecha))
                attendance_data = cursor.fetchall()
            else:
                cursor = db_manager.sqlite_connection.cursor()
                cursor.execute("""
                    SELECT nombre_completo, cargo, rol, hora_entrada, hora_salida
                    FROM empleados_local WHERE id = ? AND activo = 1
                """, (empleado_id,))
                employee_data = cursor.fetchone()

                cursor.execute("""
                    SELECT fecha, hora_registro, tipo_movimiento, estado
                    FROM registros_local 
                    WHERE empleado_id = ? AND fecha = ?
                    ORDER BY hora_registro
                """, (empleado_id, fecha.isoformat()))
                attendance_data = cursor.fetchall()

            if not employee_data:
                print(f"Empleado con ID {empleado_id} no encontrado")
                return None

            files_generated = []

            # Excel diario por empleado
            if formato in ['excel', 'both']:
                try:
                    nombre = employee_data[0]
                    filename = f"asistencia_diaria_{nombre.replace(' ', '_')}_{fecha.strftime('%Y%m%d')}.xlsx"
                    filepath = os.path.join(self.employee_reports_dir, filename)

                    df = pd.DataFrame(attendance_data, columns=['Fecha', 'Hora', 'Movimiento', 'Estado'])
                    if not df.empty:
                        df['Hora'] = pd.to_datetime(df['Hora']).dt.strftime('%H:%M:%S')

                    # Calcular horario efectivo del día
                    try:
                        he_eff, hs_eff = db_manager.obtener_horario_efectivo(empleado_id, fecha)
                    except Exception:
                        he_eff, hs_eff = str(employee_data[3])[:5], str(employee_data[4])[:5]

                    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                        summary_df = pd.DataFrame({
                            'Empleado': [nombre],
                            'Cargo': [employee_data[1]],
                            'Rol': [employee_data[2]],
                            'Horario Entrada': [he_eff],
                            'Horario Salida': [hs_eff],
                            'Fecha': [fecha.strftime('%Y-%m-%d')]
                        })
                        summary_df.to_excel(writer, sheet_name='Información', index=False)
                        df.to_excel(writer, sheet_name='Asistencias', index=False)

                        for sheet_name in writer.sheets:
                            worksheet = writer.sheets[sheet_name]
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                worksheet.column_dimensions[column_letter].width = adjusted_width

                    files_generated.append(filepath)
                except Exception as e:
                    print(f"Error generando Excel diario empleado: {e}")

            # PDF diario por empleado
            if formato in ['pdf', 'both']:
                try:
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                    from reportlab.lib import colors
                    docname = f"asistencia_diaria_{employee_data[0].replace(' ', '_')}_{fecha.strftime('%Y%m%d')}.pdf"
                    docpath = os.path.join(self.employee_reports_dir, docname)
                    doc = SimpleDocTemplate(docpath, pagesize=A4)
                    styles = getSampleStyleSheet()
                    story = []

                    title = Paragraph(f"REPORTE DIARIO - {employee_data[0].upper()}", styles['Title'])
                    subtitle = Paragraph(fecha.strftime('%d/%m/%Y'), styles['Heading2'])
                    story.append(title)
                    story.append(subtitle)
                    story.append(Spacer(1, 12))

                    # Calcular horario efectivo del día
                    try:
                        he_eff, hs_eff = db_manager.obtener_horario_efectivo(empleado_id, fecha)
                    except Exception:
                        he_eff, hs_eff = str(employee_data[3])[:5], str(employee_data[4])[:5]

                    info_data = [
                        ['Empleado:', employee_data[0]],
                        ['Cargo:', employee_data[1]],
                        ['Rol:', employee_data[2]],
                        ['Horario:', f"{he_eff} - {hs_eff}"]
                    ]
                    info_table = Table(info_data)
                    info_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ]))
                    story.append(info_table)
                    story.append(Spacer(1, 12))

                    table_data = [['Hora', 'Movimiento', 'Estado']]
                    for row in attendance_data:
                        # row: (fecha, hora, movimiento, estado)
                        if isinstance(row[1], str):
                            hora_dt = datetime.fromisoformat(row[1])
                        else:
                            hora_dt = row[1]
                        hora_str = hora_dt.strftime('%H:%M:%S')
                        table_data.append([hora_str, row[2], row[3]])

                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                    doc.build(story)
                    files_generated.append(docpath)
                except Exception as e:
                    print(f"Error generando PDF diario empleado: {e}")

            return files_generated if files_generated else None
        except Exception as e:
            print(f"Error generando reporte diario de empleado: {e}")
            return None

    def generate_employee_full_report(self, empleado_id):
        """Generar Expediente Completo (PDF) del empleado con todo el mes a mes del año actual.
        Crea un PDF con resumen e historial agrupado por mes. Devuelve lista con la ruta del archivo generado.
        """
        try:
            now = datetime.now()
            year = now.year

            # Obtener datos del empleado
            if db_manager.is_online():
                c = db_manager.pg_connection.cursor()
                c.execute("""
                    SELECT id, nombre_completo, cargo, rol, hora_entrada, hora_salida
                    FROM empleados WHERE id = %s AND activo = TRUE
                """, (empleado_id,))
                emp = c.fetchone()
                if not emp:
                    return None
                # Traer todo el año
                c.execute("""
                    SELECT fecha, hora_registro, tipo_movimiento, estado
                    FROM registros_asistencia
                    WHERE empleado_id = %s AND EXTRACT(YEAR FROM fecha) = %s
                    ORDER BY fecha, hora_registro
                """, (empleado_id, year))
                rows = c.fetchall()
            else:
                c = db_manager.sqlite_connection.cursor()
                c.execute("""
                    SELECT id, nombre_completo, cargo, rol, hora_entrada, hora_salida
                    FROM empleados_local WHERE id = ? AND activo = 1
                """, (empleado_id,))
                emp = c.fetchone()
                if not emp:
                    return None
                c.execute("""
                    SELECT fecha, hora_registro, tipo_movimiento, estado
                    FROM registros_local
                    WHERE empleado_id = ? AND substr(fecha,1,4) = ?
                    ORDER BY fecha, hora_registro
                """, (empleado_id, str(year)))
                rows = c.fetchall()

            if not rows:
                return None

            # Construir PDF
            nombre = emp[1]
            filename = f"expediente_completo_{nombre.replace(' ', '_')}_{year}.pdf"
            outpath = os.path.join(self.employee_reports_dir, filename)

            doc = SimpleDocTemplate(outpath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            story.append(Paragraph(f"EXPEDIENTE COMPLETO - {nombre.upper()}", styles['Title']))
            story.append(Paragraph(str(year), styles['Heading2']))
            story.append(Spacer(1, 12))

            info_data = [
                ['Cargo', emp[2]],
                ['Rol', emp[3]],
                ['Horario', f"{str(emp[4])[:5]} - {str(emp[5])[:5]}"]
            ]
            info_table = Table(info_data)
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 12))

            # Agrupar por mes
            by_month = {}
            for f, h, mov, est in rows:
                if isinstance(f, str):
                    fdt = datetime.fromisoformat(f)
                else:
                    fdt = f
                key = (fdt.year, fdt.month)
                by_month.setdefault(key, []).append((fdt, h, mov, est))

            month_names = [
                'Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'
            ]

            for (yy, mm) in sorted(by_month.keys()):
                story.append(Paragraph(f"{month_names[mm-1]} {yy}", styles['Heading3']))
                tdata = [['Fecha', 'Hora', 'Movimiento', 'Estado']]
                for fdt, h, mov, est in by_month[(yy, mm)]:
                    if isinstance(h, str):
                        hdt = datetime.fromisoformat(h)
                    else:
                        hdt = h
                    tdata.append([fdt.strftime('%Y-%m-%d'), hdt.strftime('%H:%M:%S'), mov, est])
                t = Table(tdata, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
                ]))
                story.append(t)
                story.append(Spacer(1, 12))

            doc.build(story)
            return [outpath]
        except Exception as e:
            print(f"Error generando expediente completo: {e}")
            return None
    
    def auto_generate_monthly_reports(self):
        """Generar automáticamente reportes mensuales de todos los empleados"""
        try:
            last_month = datetime.now().replace(day=1) - timedelta(days=1)
            year = last_month.year
            month = last_month.month
            
            # Obtener todos los empleados activos
            if db_manager.is_online():
                cursor = db_manager.pg_connection.cursor()
                cursor.execute("SELECT id, nombre_completo FROM empleados WHERE activo = TRUE")
                employees = cursor.fetchall()
            else:
                cursor = db_manager.sqlite_connection.cursor()
                cursor.execute("SELECT id, nombre_completo FROM empleados_local WHERE activo = 1")
                employees = cursor.fetchall()
            
            generated_files = []
            
            for empleado_id, nombre in employees:
                files = self.generate_employee_report(empleado_id, year, month, 'both')
                if files:
                    generated_files.extend(files)
            
            print(f"Generados {len(generated_files)} archivos de reportes automáticos")
            return generated_files
            
        except Exception as e:
            print(f"Error en generación automática: {e}")
            return []
    
    def _get_daily_data(self, fecha):
        """Obtener datos del día"""
        try:
            if db_manager.is_online():
                cursor = db_manager.pg_connection.cursor()
                cursor.execute("""
                    SELECT 
                        e.id, e.nombre_completo, e.cargo, r.hora_registro, 
                        r.tipo_movimiento, r.estado, u.nombre as ubicacion
                    FROM registros_asistencia r
                    JOIN empleados e ON r.empleado_id = e.id
                    JOIN ubicaciones u ON r.ubicacion_id = u.id
                    WHERE r.fecha = %s
                    ORDER BY r.hora_registro
                """, (fecha,))
                return cursor.fetchall()
            else:
                cursor = db_manager.sqlite_connection.cursor()
                cursor.execute("""
                    SELECT 
                        e.id, e.nombre_completo, e.cargo, r.hora_registro,
                        r.tipo_movimiento, r.estado, r.ubicacion_nombre
                    FROM registros_local r
                    JOIN empleados_local e ON r.empleado_id = e.id
                    WHERE r.fecha = ?
                    ORDER BY r.hora_registro
                """, (fecha.isoformat(),))
                return cursor.fetchall()
                
        except Exception as e:
            print(f"Error obteniendo datos diarios: {e}")
            return []
    
    def _get_monthly_data(self, year, month):
        """Obtener datos del mes"""
        try:
            if db_manager.is_online():
                cursor = db_manager.pg_connection.cursor()
                cursor.execute("""
                    SELECT 
                        e.id, e.nombre_completo, r.fecha,
                        MIN(CASE WHEN r.tipo_movimiento = 'ENTRADA' THEN r.hora_registro END) as primera_entrada,
                        MAX(CASE WHEN r.tipo_movimiento = 'SALIDA' THEN r.hora_registro END) as ultima_salida,
                        MIN(CASE WHEN r.tipo_movimiento = 'ENTRADA' THEN r.estado END) as estado_entrada,
                        MAX(CASE WHEN r.tipo_movimiento = 'SALIDA' THEN r.estado END) as estado_salida
                    FROM registros_asistencia r
                    JOIN empleados e ON r.empleado_id = e.id
                    WHERE EXTRACT(YEAR FROM r.fecha) = %s 
                    AND EXTRACT(MONTH FROM r.fecha) = %s
                    GROUP BY e.id, e.nombre_completo, r.fecha
                    ORDER BY r.fecha, e.nombre_completo
                """, (year, month))
                return cursor.fetchall()
            else:
                # Para SQLite, necesitamos una consulta diferente
                cursor = db_manager.sqlite_connection.cursor()
                cursor.execute("""
                    SELECT 
                        e.id, e.nombre_completo, r.fecha,
                        MIN(CASE WHEN r.tipo_movimiento = 'ENTRADA' THEN r.hora_registro END) as primera_entrada,
                        MAX(CASE WHEN r.tipo_movimiento = 'SALIDA' THEN r.hora_registro END) as ultima_salida,
                        MIN(CASE WHEN r.tipo_movimiento = 'ENTRADA' THEN r.estado END) as estado_entrada,
                        MAX(CASE WHEN r.tipo_movimiento = 'SALIDA' THEN r.estado END) as estado_salida
                    FROM registros_local r
                    JOIN empleados_local e ON r.empleado_id = e.id
                    WHERE substr(r.fecha, 1, 4) = ? 
                    AND substr(r.fecha, 6, 2) = ?
                    GROUP BY e.id, e.nombre_completo, r.fecha
                    ORDER BY r.fecha, e.nombre_completo
                """, (str(year), f"{month:02d}"))
                return cursor.fetchall()
                
        except Exception as e:
            print(f"Error obteniendo datos mensuales: {e}")
            return []
    
    def _get_employee_data(self, empleado_id, year, month):
        """Obtener datos específicos de un empleado"""
        try:
            # Datos del empleado
            if db_manager.is_online():
                cursor = db_manager.pg_connection.cursor()
                cursor.execute("""
                    SELECT nombre_completo, cargo, rol, hora_entrada, hora_salida
                    FROM empleados WHERE id = %s AND activo = TRUE
                """, (empleado_id,))
                employee_data = cursor.fetchone()
                
                # Datos de asistencia
                cursor.execute("""
                    SELECT fecha, hora_registro, tipo_movimiento, estado
                    FROM registros_asistencia 
                    WHERE empleado_id = %s 
                    AND EXTRACT(YEAR FROM fecha) = %s 
                    AND EXTRACT(MONTH FROM fecha) = %s
                    ORDER BY fecha, hora_registro
                """, (empleado_id, year, month))
                attendance_data = cursor.fetchall()
            else:
                cursor = db_manager.sqlite_connection.cursor()
                cursor.execute("""
                    SELECT nombre_completo, cargo, rol, hora_entrada, hora_salida
                    FROM empleados_local WHERE id = ? AND activo = 1
                """, (empleado_id,))
                employee_data = cursor.fetchone()
                
                cursor.execute("""
                    SELECT fecha, hora_registro, tipo_movimiento, estado
                    FROM registros_local 
                    WHERE empleado_id = ? 
                    AND substr(fecha, 1, 4) = ? 
                    AND substr(fecha, 6, 2) = ?
                    ORDER BY fecha, hora_registro
                """, (empleado_id, str(year), f"{month:02d}"))
                attendance_data = cursor.fetchall()
            
            return employee_data, attendance_data
            
        except Exception as e:
            print(f"Error obteniendo datos del empleado: {e}")
            return None, []
    
    def _generate_daily_excel(self, data, fecha):
        """Generar reporte diario en Excel"""
        try:
            # data esperado: (empleado_id, nombre, cargo, hora, mov, estado, ubicacion)
            df = pd.DataFrame(data, columns=[
                'EmpleadoID','Empleado', 'Cargo', 'Hora', 'Movimiento', 'Estado', 'Ubicación'
            ])
            # Formatear hora
            df['Hora'] = pd.to_datetime(df['Hora']).dt.strftime('%H:%M:%S')
            
            filename = f"reporte_diario_{fecha.strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Asistencia Diaria', index=False)

                # Formatear el archivo (centrado y colores por estado)
                from openpyxl.styles import PatternFill, Alignment, Font
                workbook = writer.book
                worksheet = writer.sheets['Asistencia Diaria']

                # Ocultar columna EmpleadoID
                worksheet.column_dimensions['A'].hidden = True

                # Ajustar ancho de columnas y centrar
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                        # Centrar todas las celdas
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Colorear filas por Estado (A_TIEMPO/TEMPRANO=verde, RETARDO/TARDE=amarillo, FALTA/NO ASISTIÓ=rojo)
                # Considerar justificaciones del día
                just_map = db_manager.obtener_justificaciones_por_fecha(fecha.isoformat())
                state_col_idx = None
                id_col_idx = None
                # Detectar encabezados
                headers = [c.value for c in next(worksheet.iter_rows(min_row=1, max_row=1))]
                for idx, h in enumerate(headers, start=1):
                    if str(h).upper() == 'ESTADO':
                        state_col_idx = idx
                    if str(h).upper() == 'EMPLEADOID':
                        id_col_idx = idx
                if state_col_idx and id_col_idx:
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        emp_id = row[id_col_idx-1].value
                        estado_val = str(row[state_col_idx-1].value or '').upper()
                        # Si RETARDO justificado => verde y añadir asterisco visual en la celda
                        if (int(emp_id) if emp_id is not None else None, 'RETARDO') in just_map and estado_val == 'RETARDO':
                            row[state_col_idx-1].value = 'RETARDO*'
                            fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                            for cell in row:
                                cell.fill = fill
                            continue
                        # Otros justificados -> neutral, sin color
                        if (int(emp_id) if emp_id is not None else None, estado_val) in just_map:
                            continue
                        fill = None
                        if estado_val in ('A_TIEMPO', 'TEMPRANO'):
                            fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                        elif estado_val in ('RETARDO', 'TARDE'):
                            fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
                        elif estado_val in ('FALTA', 'NO ASISTIO', 'NO_ASISTIO'):
                            fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                        if fill:
                            for cell in row:
                                cell.fill = fill
            
            print(f"Reporte Excel generado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error generando Excel diario: {e}")
            return None
    
    def _generate_daily_pdf(self, data, fecha):
        """Generar reporte diario en PDF"""
        try:
            filename = f"reporte_diario_{fecha.strftime('%Y%m%d')}.pdf"
            filepath = os.path.join(self.reports_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title = Paragraph(f"REPORTE DIARIO DE ASISTENCIA - {fecha.strftime('%d/%m/%Y')}", 
                            styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Tabla de datos
            table_data = [['Empleado', 'Cargo', 'Hora', 'Movimiento', 'Estado', 'Ubicación']]
            row_colors = []
            just_map = db_manager.obtener_justificaciones_por_fecha(fecha.isoformat())
            
            for row in data:
                # row puede ser de longitud 6 (sin id) o 7 (con id)
                has_id = len(row) >= 7
                if has_id:
                    emp_id, nombre, cargo, hora_val, mov, est, ubi = row
                else:
                    nombre, cargo, hora_val, mov, est, ubi = row
                    emp_id = None
                # Formatear hora
                if isinstance(hora_val, str):
                    hora_dt = datetime.fromisoformat(hora_val)
                else:
                    hora_dt = hora_val
                hora_str = hora_dt.strftime('%H:%M:%S')
                est_up = (est or '').upper()
                # Preparar estado mostrado con posible asterisco si RETARDO justificado
                estado_mostrar = est_up
                if has_id and (emp_id, 'RETARDO') in just_map and est_up == 'RETARDO':
                    estado_mostrar = 'RETARDO*'
                table_data.append([
                    str(nombre)[:20],
                    str(cargo)[:15],
                    hora_str,
                    mov,
                    estado_mostrar,
                    ubi
                ])
                # Definir color por estado si no está justificado
                if has_id and (emp_id, 'RETARDO') in just_map and est_up == 'RETARDO':
                    row_colors.append(colors.lightgreen)
                elif has_id and (emp_id, est_up) in just_map:
                    row_colors.append(colors.beige)
                else:
                    if est_up in ('A_TIEMPO', 'TEMPRANO'):
                        row_colors.append(colors.lightgreen)
                    elif est_up in ('RETARDO', 'TARDE'):
                        row_colors.append(colors.yellow)
                    elif est_up in ('FALTA', 'NO ASISTIO', 'NO_ASISTIO'):
                        row_colors.append(colors.salmon)
                    else:
                        row_colors.append(colors.beige)
            
            table = Table(table_data)
            style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]
            for ridx, bg in enumerate(row_colors, start=1):
                style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), bg))
            table.setStyle(TableStyle(style_cmds))
            
            story.append(table)
            doc.build(story)
            
            print(f"Reporte PDF generado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error generando PDF diario: {e}")
            return None
    
    def _generate_employee_excel(self, employee_data, attendance_data, year, month):
        """Generar reporte de empleado en Excel"""
        try:
            nombre = employee_data[0]
            filename = f"asistencia_{nombre.replace(' ', '_')}_{year}_{month:02d}.xlsx"
            filepath = os.path.join(self.employee_reports_dir, filename)
            
            # Crear DataFrame
            df = pd.DataFrame(attendance_data, columns=[
                'Fecha', 'Hora', 'Movimiento', 'Estado'
            ])
            
            # Formatear fechas y horas
            df['Hora'] = pd.to_datetime(df['Hora']).dt.strftime('%H:%M:%S')
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Hoja de resumen
                summary_data = {
                    'Empleado': [nombre],
                    'Cargo': [employee_data[1]],
                    'Rol': [employee_data[2]],
                    'Horario Entrada': [str(employee_data[3])[:5]],
                    'Horario Salida': [str(employee_data[4])[:5]],
                    'Mes/Año': [f"{month:02d}/{year}"]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Información', index=False)
                
                # Hoja de asistencia
                df.to_excel(writer, sheet_name='Asistencias', index=False)
                
                # Formatear
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Reporte empleado Excel generado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error generando Excel empleado: {e}")
            return None
    
    def _generate_employee_pdf(self, employee_data, attendance_data, year, month):
        """Generar reporte de empleado en PDF"""
        try:
            nombre = employee_data[0]
            filename = f"asistencia_{nombre.replace(' ', '_')}_{year}_{month:02d}.pdf"
            filepath = os.path.join(self.employee_reports_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            month_name = calendar.month_name[month]
            title = Paragraph(f"REPORTE DE ASISTENCIA - {nombre.upper()}", styles['Title'])
            subtitle = Paragraph(f"{month_name} {year}", styles['Heading2'])
            
            story.append(title)
            story.append(subtitle)
            story.append(Spacer(1, 12))
            
            # Información del empleado
            info_data = [
                ['Empleado:', nombre],
                ['Cargo:', employee_data[1]],
                ['Rol:', employee_data[2]],
                ['Horario:', f"{str(employee_data[3])[:5]} - {str(employee_data[4])[:5]}"]
            ]
            
            info_table = Table(info_data)
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 20))
            
            # Tabla de asistencias con colores por estado
            if attendance_data:
                table_data = [['Fecha', 'Hora', 'Movimiento', 'Estado']]
                row_colors = []

                for row in attendance_data:
                    if isinstance(row[1], str):
                        hora_dt = datetime.fromisoformat(row[1])
                    else:
                        hora_dt = row[1]
                    hora_str = hora_dt.strftime('%H:%M:%S')
                    est = (row[3] or '').upper()
                    table_data.append([str(row[0]), hora_str, row[2], est])
                    # Color: A_TIEMPO/TEMPRANO=verde, RETARDO/TARDE=amarillo, FALTA/NO ASISTIÓ=rojo
                    if est in ('A_TIEMPO', 'TEMPRANO'):
                        row_colors.append(colors.lightgreen)
                    elif est in ('RETARDO', 'TARDE'):
                        row_colors.append(colors.yellow)
                    elif est in ('FALTA', 'NO ASISTIO', 'NO_ASISTIO'):
                        row_colors.append(colors.salmon)
                    else:
                        row_colors.append(colors.beige)
                
                attendance_table = Table(table_data)
                style_cmds = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]
                # Colores por fila (saltando encabezado)
                for ridx, bg in enumerate(row_colors, start=1):
                    style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), bg))
                attendance_table.setStyle(TableStyle(style_cmds))
                
                story.append(attendance_table)
            else:
                story.append(Paragraph("No hay registros de asistencia para este período.", styles['Normal']))
            
            doc.build(story)
            
            print(f"Reporte empleado PDF generado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error generando PDF empleado: {e}")
            return None

    def _generate_monthly_excel(self, data, year, month):
        """Generar reporte mensual en Excel con colores por estado y centrado."""
        try:
            # data: (empleado_id, nombre, fecha, primera_entrada, ultima_salida, estado_entrada, estado_salida)
            df = pd.DataFrame(data, columns=[
                'EmpleadoID', 'Empleado', 'Fecha', 'Primera Entrada', 'Última Salida', 'Estado Entrada', 'Estado Salida'
            ])

            filename = f"reporte_mensual_{year}_{month:02d}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Asistencia Mensual', index=False)

                from openpyxl.styles import PatternFill, Alignment
                ws = writer.sheets['Asistencia Mensual']

                # Ocultar ID
                ws.column_dimensions['A'].hidden = True

                # Ajuste ancho y centrado
                for column in ws.columns:
                    max_length = 0
                    letter = column[0].column_letter
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.column_dimensions[letter].width = min(max_length + 2, 50)

                # Aplicar colores a columnas de estados si no hay justificación
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                emp_idx = headers.index('EmpleadoID') + 1
                fecha_idx = headers.index('Fecha') + 1
                ee_idx = headers.index('Estado Entrada') + 1
                es_idx = headers.index('Estado Salida') + 1

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    emp_id = row[emp_idx-1].value
                    fecha_val = row[fecha_idx-1].value
                    try:
                        fecha_iso = pd.to_datetime(fecha_val).date().isoformat()
                    except Exception:
                        fecha_iso = str(fecha_val)
                    just_map = db_manager.obtener_justificaciones_por_fecha(fecha_iso)
                    est_e = str(row[ee_idx-1].value or '').upper()
                    est_s = str(row[es_idx-1].value or '').upper()
                    def fill_for(state):
                        if state in ('A_TIEMPO', 'TEMPRANO'):
                            return PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                        if state in ('RETARDO', 'TARDE'):
                            return PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
                        if state in ('FALTA', 'NO ASISTIO', 'NO_ASISTIO'):
                            return PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                        return None
                    # RETARDO justificado: asterisco + verde
                    if (int(emp_id), 'RETARDO') in just_map and est_e == 'RETARDO':
                        row[ee_idx-1].value = 'RETARDO*'
                        row[ee_idx-1].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                    elif (int(emp_id), est_e) not in just_map:
                        f = fill_for(est_e)
                        if f:
                            row[ee_idx-1].fill = f
                    if (int(emp_id), 'RETARDO') in just_map and est_s == 'RETARDO':
                        row[es_idx-1].value = 'RETARDO*'
                        row[es_idx-1].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                    elif (int(emp_id), est_s) not in just_map:
                        f = fill_for(est_s)
                        if f:
                            row[es_idx-1].fill = f

            print(f"Reporte mensual Excel generado: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error generando Excel mensual: {e}")
            return None

    def _generate_monthly_pdf(self, data, year, month):
        """Generar reporte mensual en PDF con colores por estado y centrado."""
        try:
            filename = f"reporte_mensual_{year}_{month:02d}.pdf"
            filepath = os.path.join(self.reports_dir, filename)

            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            month_name = calendar.month_name[month]
            title = Paragraph(f"REPORTE MENSUAL DE ASISTENCIA - {month_name} {year}", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))

            table_data = [['Empleado', 'Fecha', 'Primera Entrada', 'Última Salida', 'Estado Entrada', 'Estado Salida']]
            cell_bg_cmds = []

            for idx, row in enumerate(data, start=1):
                emp_id, nombre, fecha_val, pe, us, ee, es = row
                fecha_str = str(fecha_val)
                ee_up = (ee or '').upper()
                es_up = (es or '').upper()
                # Mostrar asterisco si RETARDO justificado
                try:
                    fecha_iso = pd.to_datetime(fecha_val).date().isoformat()
                except Exception:
                    fecha_iso = fecha_str
                just_map = db_manager.obtener_justificaciones_por_fecha(fecha_iso)
                ee_show = 'RETARDO*' if ee_up == 'RETARDO' and (emp_id, 'RETARDO') in just_map else ee_up
                es_show = 'RETARDO*' if es_up == 'RETARDO' and (emp_id, 'RETARDO') in just_map else es_up
                table_data.append([
                    nombre,
                    fecha_str,
                    str(pe)[:19] if pe else '',
                    str(us)[:19] if us else '',
                    ee_show,
                    es_show
                ])
                # Justificaciones por fecha
                # ya obtenido arriba
                # Calcular colores para columnas 4 y 5 (base 0)
                def color_for(state):
                    if state in ('A_TIEMPO', 'TEMPRANO'):
                        return colors.lightgreen
                    if state in ('RETARDO', 'TARDE'):
                        return colors.yellow
                    if state in ('FALTA', 'NO ASISTIO', 'NO_ASISTIO'):
                        return colors.salmon
                    return None
                if (emp_id, 'RETARDO') in just_map and ee_up == 'RETARDO':
                    cell_bg_cmds.append(('BACKGROUND', (4, idx), (4, idx), colors.lightgreen))
                elif (emp_id, ee_up) not in just_map:
                    c = color_for(ee_up)
                    if c:
                        cell_bg_cmds.append(('BACKGROUND', (4, idx), (4, idx), c))
                if (emp_id, 'RETARDO') in just_map and es_up == 'RETARDO':
                    cell_bg_cmds.append(('BACKGROUND', (5, idx), (5, idx), colors.lightgreen))
                elif (emp_id, es_up) not in just_map:
                    c = color_for(es_up)
                    if c:
                        cell_bg_cmds.append(('BACKGROUND', (5, idx), (5, idx), c))

            table = Table(table_data)
            base_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]
            table.setStyle(TableStyle(base_style + cell_bg_cmds))

            story.append(table)
            doc.build(story)

            print(f"Reporte mensual PDF generado: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error generando PDF mensual: {e}")
            return None

# Instancia global del generador de reportes
report_generator = ReportGenerator()